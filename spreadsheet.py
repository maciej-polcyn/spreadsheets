import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
# sheet.cell(1, 1)

# print(cell.value)
# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    new_price = cell.value * 0.9
    new_price_cell = sheet.cell(row, 4)
    new_price_cell.value = new_price

new_price_column = sheet.cell(1, 4)
new_price_column.value = 'new_price'
wb.save('transactions2.xlsx')